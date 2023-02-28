import os
import sys

from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

import requests

import ssl
ssl._create_default_https_context = ssl._create_unverified_context

from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver import EdgeOptions

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from selenium.common.exceptions import NoSuchElementException, \
    ElementNotVisibleException, TimeoutException, StaleElementReferenceException

import pdfkit

import re

from pytesseract import *
import cv2

import openpyxl

import win32com.client as win32


def check_element_by_xpath(xpath):
    r = False
    try:
        driver.find_element(By.XPATH, xpath)
        r = True
    except (StaleElementReferenceException, NoSuchElementException) as e:
        r = False
    return r


def retry_find_element_by_xpath(xpath, nret):
    result = False
    attempts = 0

    while attempts < nret:
        try:
            driver.find_element(By.XPATH, xpath)
            result = True
            break
        except (StaleElementReferenceException, NoSuchElementException) as e:
            pass
        attempts += 1
    return result


def ngen():
    num = 0
    while True:
        yield num
        num += 1


def remove_files(path, incl_parent='n'):
    if os.path.isdir(path):
        for root, dirs, files in os.walk(os.path.abspath(path), topdown=False):
            for f in files:
                os.remove(os.path.join(root, f))
            for d in dirs:
                os.rmdir(os.path.join(root, d))
        if incl_parent == 'y':
            os.rmdir(path)


def atoi(text):
    return int(text) if text.isdigit() else text


def natural_keys(text):
    return [ atoi(c) for c in re.split(r'(\d+)', text) ]


def get_number_from_s(s):
    if '.' in s:
        return float(s)
    else:
        return int(s)


def weird_division(n, d):
    return n / d if d else 0


def get_dt(month):
    dt = datetime.now()
    st = dt - relativedelta(months=month)
    st = st.replace(day=1, hour=0, minute=0, second=0)

    et = st
    et = et + relativedelta(months=1) - timedelta(days=1)
    et = et.replace(hour=23, minute=59, second=59)

    return st, et


w_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
parent_images_dir = os.path.join(w_dir, 'images')
output_dir = os.path.join(w_dir, 'output')

drivers_dir = os.path.join(w_dir, 'drivers')
driver_path = os.path.join(drivers_dir, 'edgedriver_win64', 'msedgedriver.exe')

remove_files(parent_images_dir, 'n')
remove_files(output_dir, 'n')

service = EdgeService(driver_path)
options = EdgeOptions()
options.add_argument('--ignore-ssl-errors=yes')
options.add_argument('--ignore-certificate-errors')
#options.add_experimental_option("detach", True)
driver = webdriver.Edge(service=service, options=options)

driver.get('https://192.168.100.20/zabbix/zabbix.php?action=dashboard.view')

WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//a[@class='top-nav-signout']"))).click()
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[@id='name']"))).send_keys('Admin')
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[@id='password']"))).send_keys('Admin')
WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//button[@id='enter']"))).click()

WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//a[text()='Screens']"))).click()

raport_map = {
    'Capacity Report (Razvan - Part 1)': 48,
    'Capacity Report (Razvan - Part 2)': 38,
    'Capacity Report (Razvan - Part 3)': 34
}

user_agent = dict({"User-agent": driver.execute_script("return navigator.userAgent;")})

c = driver.get_cookies()
cookies = {}
for entry in c:
    cookies.update({entry["name"]: entry["value"]})

dts = [(get_dt(1)), (get_dt(2))]

image_dirs = []
for dt in dts:
    field_st_ts = f"{dt[0].strftime('%Y-%m-%d %H:%M:%S')}"
    field_et_ts = f"{dt[1].strftime('%Y-%m-%d %H:%M:%S')}"
    ts = f"{dt[0].strftime('%Y-%m-%d')}-{dt[1].strftime('%Y-%m-%d')}"

    images_dir = os.path.join(parent_images_dir, ts)
    image_dirs.append(images_dir)
    os.mkdir(images_dir)


    image_name = ngen()
    for r_name, r_item_count in raport_map.items():
        if len(driver.find_elements(By.XPATH, "//a[text()='All screens']")) > 0:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//a[text()='All screens']"))).click()

        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, f"/html/body/main/form/table/tbody/tr/td/a[text()='{r_name}']"))).click()

        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='from']"))).clear()
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='from']"))).send_keys(field_st_ts)
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='to']"))).clear()
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='to']"))).send_keys(field_et_ts)
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//button[@id='apply']"))).click()

        num = ngen()

        cur_count = 0
        for try_count in num:
            if not check_element_by_xpath(f"/html/body/main/div/table/tbody/tr[{try_count}]/td/div/a/img"):
                continue
            image = driver.find_element(By.XPATH, f"/html/body/main/div/table/tbody/tr[{try_count}]/td/div/a/img")
            response = requests.get(image.get_attribute("src"), stream=True, headers=user_agent, cookies=cookies, verify=False)
            with open(os.path.join(images_dir, f"{str(next(image_name))}.png"), 'wb') as f:
                f.write(response.content)
            cur_count += 1
            if cur_count == r_item_count:
                break

st, et = get_dt(1)
ts = f"{st.strftime('%Y-%m-%d')}-{et.strftime('%Y-%m-%d')}"
output_basename = f"capacity_report_{ts}"


config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
options = {
    'dpi': 400,
    'page-size': 'A2',
    'margin-top': '0.25in',
    'margin-right': '0.25in',
    'margin-bottom': '0.25in',
    'margin-left': '0.25in',
    'encoding': "UTF-8",
    'custom-header': [
        ('Accept-Encoding', 'gzip')
    ],
    'no-outline': None,
    'enable-local-file-access': None
}

html_file = os.path.join(output_dir, f"{output_basename}.html")
pdf_file = os.path.join(output_dir, f"{output_basename}.pdf")

with open(html_file, 'w') as f:
    f.write('<html><head></head><body>')

    for root, dirs, files in os.walk(image_dirs[0]):
        for filename in sorted(files, key=natural_keys):
            f.write('<img src=\"' + os.path.join(root, filename) + '\">\n')

    f.write('</body></html>')
    f.close()

pdfkit.from_file(html_file, pdf_file, configuration=config, options=options)


corrections = [
    (re.compile(r'\bOevent\b'), '0 event'),
    (re.compile(r'\bObyte\b'), '0 byte'),
    (re.compile(r'\[ava\]'), '[avg]'),
    (re.compile(r'\bQevent\b'), '0 event'),
    (re.compile(r'\bO events\b'), '0 events'),
    (re.compile(r'\boO\b'), '0'),
    (re.compile(r'\b3vg\b'), 'avg'),
    (re.compile(r'\bOeventis\b'), '0 events'),
    (re.compile(r'\bSOKbit\b'), '0 Kbit'),
    (re.compile(r'\bO bit\b'), '0 bit'),
    (re.compile(r'\bG6B\b'), 'GB'),
    (re.compile(r'\bSmin\b'), '5 min'),
    (re.compile(r'(\bfavg[\)}]?\b)'), '[avg]'),
    (re.compile(r'(\[0S\])'), '[OS]'),
    (re.compile(r'(o0am|0am)'), 'oam'),
    (re.compile(r'(Soam)'), '5oam'),
    (re.compile(r'(°)'), ''),
    (re.compile(r'(?<=\.)(tde)'), 'tdc'),
    (re.compile(r'(?<=\d)(,|\]|\\n\@|\\n\©)'), ''),
    (re.compile(r'(\bALBEEACTRI\b)'), 'ALBEEACTR1'),
    (re.compile(r'(\bALBEEACORI\b)'), 'ALBEEACOR1'),
    (re.compile(r'(\bSLTEEAGTPUL\b)'), 'SLTEEAGTPU1'),
    (re.compile(r'(\bALBEEAGTPUL\b)'), 'ALBEEAGTPU1'),
    (re.compile(r'(\bKDEEAGTPUL1\b)'), 'KDEEAGTPU1'),
    (re.compile(r'(\bBOAEEAGTPUL\b)'), 'BOAEEAGTPU1'),
    (re.compile(r'ALBEEAARKS5oam'), 'ALBEEAARK5oam'),
    (re.compile(r'ALBEEACORI1'), 'ALBEEACOR1'),
    (re.compile(r'ALBEEACTRI1'), 'ALBEEACTR1'),
    (re.compile(r'ALBEEAGPEHL1'), 'ALBEEAGPEH1'),
    (re.compile(r'BOAEEAGTPUL1'), 'BOAEEAGTPU1'),
    (re.compile(r'SLTEEAGTPUL1'), 'SLTEEAGTPU1'),
    (re.compile(r'ALBEEAIMPS5'), 'ALBEEAIMP5'),
    (re.compile(r'ALBEEAGPEHL1'), 'ALBEEAGPEH1'),
    (re.compile(r'ALBEEAGPEHL'), 'ALBEEAGPEH1'),
    (re.compile(r'ALBEEAGPEHL1'), 'ALBEEAGPEH1'),
    (re.compile(r'SLTEEAGTPUS3'), 'SLTEEAGTPU3'),
    (re.compile(r'ALBEEAGTPUL1'), 'ALBEEAGTPU1'),
    (re.compile(r'ALBEEACOR3 S'), 'ALBEEACOR3'),
    (re.compile(r'(?:[A-Z]{2})(?:—|_|—_|_—|__|——|-)+'), ''),
    (re.compile(r'(?!.*oam|.*\dPP)(\d*\.?\d+)([a-zA-Z%\/]+)'), fr'\1 \2')
]

dict_ = {}
pytesseract.tesseract_cmd = fr'''C:\Users\{os.getlogin()}\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'''
for i in dts:
    ts = f"{i[0].strftime('%Y-%m-%d')}-{i[1].strftime('%Y-%m-%d')}"
    images_dir = os.path.join(parent_images_dir, ts)
    for root, dirs, files in os.walk(images_dir):
        for filename in sorted(files, key=natural_keys):
            image = os.path.join(root, filename)
            image = cv2.imread(image)
            height, width = image.shape[:2]
            res_image = cv2.resize(image, (2*width, 2*height), interpolation=cv2.INTER_CUBIC)
            data = image_to_string(res_image, config='--psm 6')
            for j in corrections:
                data = j[0].sub(j[1], data)
            key_rx = re.compile(r'([a-zA-Z0-9\-\.]*:.*[a-zA-Z0-9\-]+)')
            try:
                key = re.findall(key_rx, data)[0]
            except IndexError:
                key = re.findall(key_rx, data)
            key = key.lstrip(': ')

            values = re.findall(r"(\b\w{3,}\b.*?\[avg\])\D+(\d*\.?\d+\D+){3}(\d*\.?\d+\D*?)\n", data)

            dict_.setdefault(key, {}).update({ts: values})

print(dict_)

wb_file = os.path.join(output_dir, f"{output_basename}.xlsx")
wb = openpyxl.Workbook()
ws = wb.active

ws.cell(row=1, column=3).value = 'Current'
ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
ws.cell(row=1, column=5).value = 'Previous'
ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)

ws.append(['Name', 'Item', 'Average', 'Maximum', 'Average', 'Maximum', 'Average Difference', 'Maximum Difference', 'Unit'])

ct = f"{dts[0][0].strftime('%Y-%m-%d')}-{dts[0][1].strftime('%Y-%m-%d')}"
pt = f"{dts[1][0].strftime('%Y-%m-%d')}-{dts[1][1].strftime('%Y-%m-%d')}"

for k, v in dict_.items():
    x = get_number_from_s(dict_[k][ct][0][1].split(' ')[0].rstrip('.'))
    y = get_number_from_s(dict_[k][pt][0][1].split(' ')[0].rstrip('.'))
    avg_diff = (weird_division(x - y, y)) * 100

    x = get_number_from_s(dict_[k][ct][0][2].split(' ')[0].rstrip('.'))
    y = get_number_from_s(dict_[k][pt][0][2].split(' ')[0].rstrip('.'))
    max_diff = (weird_division(x - y, y)) * 100

    unit = dict_[k][ct][0][1].split(' ')[1] if len(dict_[k][ct][0][1].split(' ')) > 1 else dict_[k][ct][0][1]

    result = [
        k,
        dict_[k][ct][0][0],
        dict_[k][ct][0][1].split(' ')[0],
        dict_[k][pt][0][1].split(' ')[0],
        dict_[k][ct][0][2].split(' ')[0],
        dict_[k][pt][0][2].split(' ')[0],
        "%.2f"% avg_diff,
        "%.2f"% max_diff,
        unit
    ]

    ws.append(result)

wb.save(wb_file)


html = f'''
<html>
    <body>
        <p>
            <p>Hello, </p>

            <p>Please find the capacity report for the range of <b>{ct}</b>, 
            compared to the previous raport with the range <b>{pt}</b>. </p>

            <p>Regards, <p>
            <p>Someone <p>
        </p>
    </body>
</html>
'''

outlook = win32.Dispatch('outlook.application')
mail = outlook.CreateItem(0)
mail.To = 'exampleuser1@example.com'
mail.CC = 'exampleuser2@example.com'
mail.Subject = f"Capacity Report {ct}"
mail.HTMLBody = html
mail.Attachments.Add(pdf_file)
mail.Attachments.Add(wb_file)
mail.Send()
